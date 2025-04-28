import os
import pandas as pd
import portalocker
from openpyxl import load_workbook

# Caminho para o Excel
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
FILENAME = next(f for f in os.listdir(DATA_DIR) if f.lower().endswith('.xlsx'))
DATA_PATH = os.path.join(DATA_DIR, FILENAME)
SHEET_NAME = "Criação Ambientes e Equipamento"

def _find_header(ws):
    """
    Varre ws linha a linha até achar 'Sequencial' na coluna A.
    Retorna (header_row_number, header_values_list).
    """
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first = row[0]
        if isinstance(first, str) and first.strip().lower() == 'sequencial':
            header = [cell.strip() if isinstance(cell, str) else cell
                      for cell in row]
            print(f"[utils] Header found at Excel row {i}: {header}")
            return i, header
    raise ValueError("Não encontrei a linha de cabeçalho com 'Sequencial'.")

def read_data():
    """
    Abre o workbook, detecta a linha de cabeçalho dinamicamente e
    constrói um DataFrame com ela e as linhas abaixo.
    """
    wb = load_workbook(DATA_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    header_row, header = _find_header(ws)

    data = []
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        if all(cell is None for cell in row):
            continue
        data.append(list(row))

    df = pd.DataFrame(data, columns=header)
    df.columns = df.columns.str.replace('\n', ' ', regex=False).str.strip()
    return df

def write_data(df):
    """
    Abre o workbook, detecta novamente a linha de cabeçalho e
    atualiza célula-a-célula apenas o corpo de dados, preservando
    o cabeçalho e demais abas.
    """
    with open(DATA_PATH, 'r+b') as lockfile:
        portalocker.lock(lockfile, portalocker.LockFlags.EXCLUSIVE)
        wb = load_workbook(DATA_PATH)
        ws = wb[SHEET_NAME]

        header_row, header = _find_header(ws)
        col_to_idx = {header[j]: j+1 for j in range(len(header))}

        for i, row in df.iterrows():
            excel_row = header_row + 1 + i
            for col, val in row.items():
                idx = col_to_idx.get(col)
                if idx:
                    ws.cell(row=excel_row, column=idx).value = val

        wb.save(DATA_PATH)
        portalocker.unlock(lockfile)